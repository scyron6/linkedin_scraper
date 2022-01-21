from openpyxl import load_workbook, Workbook
from scrape import get_contacts

filepath = 'contacts.xlsx'

try:
    wb = load_workbook(data_file, data_only=True)
    ws = wb['Sheet']
except:
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Email'])

# Delete all Rows other than Headers
while (ws.max_row > 1):
    ws.delete_rows(2)

# Scrape and Append Rows
for contact in get_contacts():
    ws.append([contact[0],contact[1]])

wb.save(filename=filepath)
