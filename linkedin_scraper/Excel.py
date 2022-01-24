from openpyxl import load_workbook, Workbook

class Excel:
    def __init__(self, filepath):
        self.filepath = filepath

        try:
            self.wb = load_workbook(self.filepath, data_only=True)
        except:
            self.wb = Workbook()
            ws = self.wb.active
            ws.append(['Name', 'Email'])
            self.wb.save(filename=self.filepath)

    def clear_sheet(self):
        ws = self.wb.active
        while (ws.max_row > 1):
            ws.delete_rows(2)

        self.wb.save(filename=self.filepath)

    def write_contacts(self, contacts):
        ws = self.wb.active
        for contact in contacts:
            ws.append([contact[0], contact[1]])

        self.wb.save(filename=self.filepath)
