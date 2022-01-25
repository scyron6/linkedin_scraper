from openpyxl import load_workbook, Workbook

from .error import InvalidWorkbookError

class Excel:
    def __init__(self, filepath):
        self.filepath = filepath

        try:
            self.wb = load_workbook(self.filepath, data_only=True)
            self.validate_excel()
        except:
            self.wb = Workbook()
            ws = self.wb.active
            ws.append(['Name', 'Email'])
            self.wb.save(filename=self.filepath)

    def clear_sheet(self):
        ws = self.wb.active
        while (ws.max_row > 1):
            ws.delete_rows(2)

        self.wb.save(filename=self.filepath)

    def write_contacts(self, contacts):
        ws = self.wb.active
        for contact in contacts:
            ws.append([contact[0], contact[1]])

        self.wb.save(filename=self.filepath)

    def validate_excel(self):
        ws = self.wb.active
        
        # Validate excel
        if ws.max_column > 2:
            raise InvalidWorkbookError

        if ws.cell(row=1, column=1).value != 'Name':
            raise InvalidWorkbookError

        if ws.cell(row=1, column=2).value != 'Email':
            raise InvalidWorkbookError