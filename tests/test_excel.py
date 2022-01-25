import os

from openpyxl import load_workbook
from linkedin_scraper.excel import Excel

filename = 'newfile.xlsx'

# When given filename that doesn't exist, Excel creates a new one
def test_create_new_file():
    workbook = Excel(filename)
    assert os.path.exists(filename)
    os.remove(filename)

# When creating new file, Excel creates header row
def test_create_header_row():
    workbook = Excel(filename)
    ws = workbook.wb.active
    assert ws.cell(row=1, column=1).value == 'Name'
    assert ws.cell(row=1, column=2).value == 'Email'
    assert ws.max_row == 1
    assert ws.max_column == 2
    os.remove(filename)

# Clear ws deletes all rows except header row
def test_clear_ws_deletes_all_but_first_row():
    workbook = Excel(filename)
    ws = workbook.wb.active
    for x in range(5):
        ws.append(['sample_name', x])

    assert ws.max_row == 6

    workbook.clear_sheet()
    assert ws.cell(row=1, column=1).value == 'Name'
    assert ws.cell(row=1, column=2).value == 'Email'
    assert ws.max_row == 1
    assert ws.max_column == 2
    os.remove(filename)

# write_contacts successfully writes one row
def test_write_contacts_writes_one_contact():
    workbook = Excel(filename)
    ws = workbook.wb.active
    contacts = [('stephen', 'email@gmail.com')]

    workbook.write_contacts(contacts)

    assert ws.max_row == 2
    assert ws.max_column == 2
    assert ws.cell(row=2, column=1).value == 'stephen'
    assert ws.cell(row=2, column=2).value == 'email@gmail.com'

    os.remove(filename)

# write_contacts successfully writes two rows
def test_write_contacts_writes_two_contacts():
    workbook = Excel(filename)
    ws = workbook.wb.active
    contacts = [('stephen', 'email@gmail.com'), ('mark', 'mark@gmail.com')]

    workbook.write_contacts(contacts)

    assert ws.max_row == 3
    assert ws.max_column == 2
    assert ws.cell(row=2, column=1).value == 'stephen'
    assert ws.cell(row=2, column=2).value == 'email@gmail.com'
    assert ws.cell(row=3, column=1).value == 'mark'
    assert ws.cell(row=3, column=2).value == 'mark@gmail.com'

    os.remove(filename)